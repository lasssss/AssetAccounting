import os, sys, uuid, io, zipfile, tempfile, shutil
from datetime import date, datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, abort, send_from_directory, send_file
from config import Config, app_dir, resources_dir
from models import db, Registry, Asset, Category, Movement, StatusLog, CardPhoto, AssetPhoto, CardListTitle, CardListItem, get_registry_session, seed_categories, Asset as AssetModel
from sqlalchemy import or_
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__, template_folder=os.path.join(resources_dir(), 'templates'),
            static_folder=os.path.join(resources_dir(), 'static'))
app.config.from_object(Config)
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(app_dir(), 'instance', 'assets.db')}"
db.init_app(app)

# ─── Helpers ──────────────────────────────────────────────────────────

def get_registry(registry_id):
    return Registry.query.get_or_404(registry_id)

def open_registry(registry_id):
    reg = get_registry(registry_id)
    return reg, get_registry_session(os.path.join(app_dir(), reg.db_file))


FIELD_MAP = {
    'инв. №': 'inventory_number', 'инвентарный номер': 'inventory_number',
    'наименование': 'name', 'подразделение': 'department', 'шифр': 'cipher',
    'стоимость': 'initial_cost', 'первоначальная стоимость': 'initial_cost',
    'срок пол.исп.': 'useful_life_months', 'спи': 'useful_life_months',
    'дата поступления': 'start_date', 'дата ввода': 'start_date', 'дата': 'start_date',
    's/n': 'serial_number', 'тип': 'equipment_type', 'место установки': 'location',
    'место': 'location', 'признак': 'indicator', 'мол': 'responsible_person',
    'ответственный': 'responsible_person', 'примечание': 'notes', 'примечания': 'notes',
    'инф.': 'info', 'кол-во': 'quantity', 'количество': 'quantity',
    'ед.изм.': 'unit', 'единица измерения': 'unit',
    'карточка': 'card', 'объект установки': 'card',
    'дата, с которой не используется (в зип)': 'inactive_date',
    'дата вывода зип': 'inactive_date', 'дата вывода': 'inactive_date',
    'дата начала работы (из зипа)': 'resume_date',
    'дата начала работы зип': 'resume_date', 'дата работы зип': 'resume_date',
    'дата начала работы зипа': 'resume_date', 'дата работы зипа': 'resume_date',
    'ликвидационная стоимость': 'salvage_value', 'ликвидационная': 'salvage_value',
}

SKIP_PHRASES = ['остаточная стоимость']


def normalize(s):
    return s.strip().lower().replace('\xa0', ' ').replace('ё', 'е')


def resolve_category(sess, name):
    if not name or not name.strip():
        cat = sess.query(Category).filter_by(name='Прочие основные средства').first()
        return cat.id if cat else sess.query(Category).first().id
    cat = sess.query(Category).filter_by(name=name.strip()).first()
    if cat:
        return cat.id
    cat = Category(name=name.strip())
    sess.add(cat)
    sess.flush()
    return cat.id


def guess_columns(headers):
    result = {}
    for i, h in enumerate(headers):
        hl = normalize(str(h))
        if any(s in hl for s in SKIP_PHRASES):
            continue
        best, best_len = None, 0
        for key, field in FIELD_MAP.items():
            nk = normalize(key)
            if nk == hl:
                best = field
                break
            if nk in hl or hl in nk:
                if len(nk) > best_len:
                    best = field
                    best_len = len(nk)
        if best:
            result[best] = i
    return result


def parse_float(v):
    try:
        return float(str(v).replace(' ', '').replace('\xa0', '').replace(',', '.'))
    except:
        return 0.0


def parse_date_val(v):
    if pd.isna(v):
        return date.today()
    try:
        return pd.to_datetime(v).date()
    except:
        try:
            return date.fromisoformat(str(v))
        except:
            return date.today()


# ─── Routes ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    registries = Registry.query.order_by(Registry.created_at.desc()).all()
    return render_template('index.html', registries=registries)


@app.route('/export-db')
def export_db():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        base = app_dir()
        for folder in ['instance', 'databases']:
            fpath = os.path.join(base, folder)
            if os.path.isdir(fpath):
                for root, dirs, files in os.walk(fpath):
                    for fn in files:
                        full = os.path.join(root, fn)
                        arcname = os.path.relpath(full, base)
                        zf.write(full, arcname)
        uploads = os.path.join(base, 'uploads', 'photos')
        if os.path.isdir(uploads):
            for fn in os.listdir(uploads):
                full = os.path.join(uploads, fn)
                if os.path.isfile(full):
                    zf.write(full, os.path.join('uploads', 'photos', fn))
    buf.seek(0)
    return send_file(buf, download_name=f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip',
                     as_attachment=True)


@app.route('/import-db', methods=['POST'])
def import_db():
    if 'file' not in request.files:
        flash('Файл не выбран', 'danger')
        return redirect(url_for('index'))
    f = request.files['file']
    if f.filename == '':
        flash('Файл не выбран', 'danger')
        return redirect(url_for('index'))
    if not f.filename.endswith('.zip'):
        flash('Требуется .zip архив', 'danger')
        return redirect(url_for('index'))
    base = app_dir()
    with tempfile.TemporaryDirectory() as tmp:
        zippath = os.path.join(tmp, 'import.zip')
        f.save(zippath)
        with zipfile.ZipFile(zippath, 'r') as zf:
            zf.extractall(tmp)
        for folder in ['instance', 'databases', 'uploads']:
            src = os.path.join(tmp, folder)
            if os.path.isdir(src):
                dst = os.path.join(base, folder)
                for root, dirs, files in os.walk(src):
                    rel = os.path.relpath(root, src)
                    target = os.path.join(dst, rel) if rel != '.' else dst
                    os.makedirs(target, exist_ok=True)
                    for fn in files:
                        shutil.copy2(os.path.join(root, fn), os.path.join(target, fn))
    flash('База данных восстановлена. Перезапустите приложение.', 'success')
    return redirect(url_for('index'))


@app.route('/registry/new', methods=['GET', 'POST'])
def registry_create():
    if request.method == 'POST':
        name = request.form['name'].strip()
        if not name:
            flash('Введите название реестра', 'danger')
            return render_template('registry_form.html')
        filename = f'registry_{uuid.uuid4().hex[:12]}.db'
        db_path = os.path.join(app_dir(), 'databases', filename)
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        reg = Registry(name=name, description=request.form.get('description', ''),
                       db_file=os.path.join('databases', filename))
        db.session.add(reg)
        db.session.commit()
        sess = get_registry_session(db_path)
        seed_categories(sess)
        sess.close()
        flash(f'Реестр "{name}" создан', 'success')
        return redirect(url_for('registry_view', registry_id=reg.id))
    return render_template('registry_form.html')


@app.route('/registry/<int:registry_id>')
def registry_view(registry_id):
    reg, sess = open_registry(registry_id)
    q = sess.query(Asset).filter(Asset.is_deleted != True)

    search = request.args.get('search', '').strip()
    filter_card = request.args.get('card', '').strip()
    filter_type = request.args.get('type', '').strip()
    filter_cat = request.args.get('category', '').strip()
    filter_status = request.args.get('status', '').strip()
    zero_residual = request.args.get('zero_residual') == '1'

    if search:
        like = f'%{search}%'
        q = q.filter(or_(
            Asset.inventory_number.like(like),
            Asset.name.like(like),
            Asset.department.like(like),
            Asset.location.like(like),
            Asset.responsible_person.like(like),
            Asset.cipher.like(like),
        ))
    if filter_card:
        q = q.filter(Asset.card == filter_card)
    if filter_type:
        q = q.filter(Asset.equipment_type == filter_type)
    if filter_cat:
        q = q.filter(Asset.category_id == int(filter_cat))
    if filter_status:
        q = q.filter(Asset.status == filter_status)

    assets = q.order_by(Asset.inventory_number).all()
    if zero_residual:
        assets = [a for a in assets if a.salvage_value <= 0]
    categories = sess.query(Category).order_by(Category.name).all()

    base_filter = (Asset.is_deleted != True)
    cards = [r[0] for r in sess.query(Asset.card).filter(Asset.card.isnot(None), Asset.card != '', base_filter).distinct().order_by(Asset.card).all()]
    types = [r[0] for r in sess.query(Asset.equipment_type).filter(Asset.equipment_type.isnot(None), Asset.equipment_type != '', base_filter).distinct().order_by(Asset.equipment_type).all()]
    statuses = [r[0] for r in sess.query(Asset.status).filter(Asset.status.isnot(None), base_filter).distinct().order_by(Asset.status).all()]

    stats = {
        'total': len(assets),
        'active': sum(1 for a in assets if a.status == 'active'),
        'disposed': sum(1 for a in assets if a.status == 'disposed'),
        'total_cost': sum(a.initial_cost for a in assets),
    }
    html = render_template('registry_view.html', registry=reg, assets=assets,
                           categories=categories, stats=stats,
                           cards=cards, types=types, statuses=statuses,
                            filter_card=filter_card, filter_type=filter_type,
                            filter_cat=filter_cat, filter_status=filter_status,
                            search=search, zero_residual=zero_residual,
                            query_params=request.query_string.decode('utf-8'))
    sess.close()
    return html


@app.route('/registry/<int:registry_id>/export-xlsx')
def registry_export(registry_id):
    reg, sess = open_registry(registry_id)
    q = sess.query(Asset).filter(Asset.is_deleted != True)

    search = request.args.get('search', '').strip()
    filter_card = request.args.get('card', '').strip()
    filter_type = request.args.get('type', '').strip()
    filter_cat = request.args.get('category', '').strip()
    filter_status = request.args.get('status', '').strip()
    zero_residual = request.args.get('zero_residual') == '1'

    if search:
        like = f'%{search}%'
        q = q.filter(or_(
            Asset.inventory_number.like(like), Asset.name.like(like),
            Asset.department.like(like), Asset.location.like(like),
            Asset.responsible_person.like(like), Asset.cipher.like(like),
        ))
    if filter_card:
        q = q.filter(Asset.card == filter_card)
    if filter_type:
        q = q.filter(Asset.equipment_type == filter_type)
    if filter_cat:
        q = q.filter(Asset.category_id == int(filter_cat))
    if filter_status:
        q = q.filter(Asset.status == filter_status)

    assets = q.order_by(Asset.inventory_number).all()
    if zero_residual:
        assets = [a for a in assets if a.salvage_value <= 0]

    rows = []
    for a in assets:
        rows.append({
            'Инв. №': a.inventory_number,
            'Наименование': a.name,
            'Остат. стоимость': a.salvage_value,
            'Подразделение': a.department or '',
            'Шифр': a.cipher or '',
            'Категория': a.category_rel.name if a.category_rel else '',
            'Стоимость': a.initial_cost,
            'Срок пол.исп. (мес)': a.useful_life_months,
            'Дата ввода': a.start_date.isoformat() if a.start_date else '',
            'Место установки': a.location or '',
            'МОЛ': a.responsible_person or '',
            'Карточка': a.card or '',
            'Признак': a.indicator or '',
            'S/N': a.serial_number or '',
            'Тип': a.equipment_type or '',
            'Кол-во': a.quantity or 1,
            'Ед.изм.': a.unit or '',
            'Дата вывода (ЗИП)': a.inactive_date.isoformat() if a.inactive_date else '',
            'Дата работы (ЗИПа)': a.resume_date.isoformat() if a.resume_date else '',
            'Статус': a.status_display,
            'Дата статуса': a.status_date.isoformat() if a.status_date else '',
            'Примечания': a.notes or '',
        })
    sess.close()
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ОС')
        ws = writer.sheets['ОС']
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(
                sum(1.7 if ord(c) > 127 else 1 for c in str(cell.value or ''))
                for cell in col_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 2, 42), 8)
    buf.seek(0)
    safe = reg.name.replace('/', '_').replace('\\', '_').replace(':', '_')
    return send_file(buf, download_name=f'{safe}.xlsx', as_attachment=True)


@app.route('/registry/<int:registry_id>/cards')
def registry_cards(registry_id):
    reg, sess = open_registry(registry_id)
    rows = sess.query(Asset.card, db.func.count(Asset.id)).filter(
        Asset.card.isnot(None), Asset.card != '', Asset.is_deleted != True
    ).group_by(Asset.card).order_by(Asset.card).all()
    cards = [{'name': r[0], 'count': r[1]} for r in rows]
    sess.close()
    return render_template('cards.html', registry=reg, cards=cards)


@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>')
def registry_cards_api(registry_id, card_name):
    reg, sess = open_registry(registry_id)
    assets = sess.query(Asset).filter(Asset.card == card_name, Asset.is_deleted != True).order_by(Asset.inventory_number).all()
    data = []
    for a in assets:
        data.append({
            'id': a.id,
            'inventory_number': a.inventory_number,
            'name': a.name,
            'status': a.status,
            'status_display': a.status_display,
            'location': a.location or '',
            'responsible_person': a.responsible_person or '',
            'initial_cost': a.initial_cost,
            'card': a.card or '',
        })
    photos = sess.query(CardPhoto).filter(CardPhoto.card_name == card_name).order_by(CardPhoto.uploaded_at.desc()).all()
    photo_list = [{'id': p.id, 'filename': p.filename, 'original_name': p.original_filename} for p in photos]
    sess.close()
    return {'assets': data, 'photos': photo_list}


UPLOAD_DIR = os.path.join(app_dir(), 'uploads', 'photos')


@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>/upload', methods=['POST'])
def card_photo_upload(registry_id, card_name):
    reg, sess = open_registry(registry_id)
    if 'photo' not in request.files:
        sess.close()
        return {'error': 'Нет файла'}, 400
    file = request.files['photo']
    if not file.filename:
        sess.close()
        return {'error': 'Файл не выбран'}, 400
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename)[1] or '.jpg'
    stored_name = f'{uuid.uuid4().hex}{ext}'
    file.save(os.path.join(UPLOAD_DIR, stored_name))
    photo = CardPhoto(card_name=card_name, filename=stored_name, original_filename=file.filename)
    sess.add(photo)
    sess.commit()
    sess.close()
    return {'ok': True, 'filename': stored_name, 'original_name': file.filename}


@app.route('/uploads/photos/<filename>')
def card_photo_serve(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# ─── Card checklist API ────────────────────────────────────────────────

@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>/list')
def card_list_get(registry_id, card_name):
    reg, sess = open_registry(registry_id)
    title_row = sess.query(CardListTitle).filter_by(card_name=card_name).first()
    items = sess.query(CardListItem).filter_by(card_name=card_name).order_by(CardListItem.position).all()
    sess.close()
    return {
        'title': title_row.title if title_row else '',
        'items': [{'id': i.id, 'text': i.text, 'position': i.position} for i in items],
    }


@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>/list/title', methods=['PUT'])
def card_list_title(registry_id, card_name):
    reg, sess = open_registry(registry_id)
    data = request.get_json(force=True)
    title_row = sess.query(CardListTitle).filter_by(card_name=card_name).first()
    if not title_row:
        title_row = CardListTitle(card_name=card_name, title=data.get('title', ''))
        sess.add(title_row)
    else:
        title_row.title = data.get('title', '')
    sess.commit()
    sess.close()
    return {'ok': True}


@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>/list/items', methods=['POST'])
def card_list_add(registry_id, card_name):
    reg, sess = open_registry(registry_id)
    data = request.get_json(force=True)
    max_pos = sess.query(db.func.max(CardListItem.position)).filter_by(card_name=card_name).scalar() or 0
    item = CardListItem(card_name=card_name, text=data.get('text', ''), position=max_pos + 1)
    sess.add(item)
    sess.commit()
    new_id = item.id
    sess.close()
    return {'ok': True, 'id': new_id, 'position': max_pos + 1}


@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>/list/items/<int:item_id>', methods=['DELETE'])
def card_list_delete(registry_id, card_name, item_id):
    reg, sess = open_registry(registry_id)
    item = sess.query(CardListItem).filter_by(id=item_id, card_name=card_name).first()
    if item:
        sess.delete(item)
        sess.commit()
    sess.close()
    return {'ok': True}


@app.route('/registry/<int:registry_id>/api/cards/<path:card_name>/export')
def card_export(registry_id, card_name):
    reg, sess = open_registry(registry_id)
    assets = sess.query(Asset).filter(Asset.card == card_name, Asset.is_deleted != True).order_by(Asset.inventory_number).all()
    rows = []
    for a in assets:
        rows.append({
            'Инв. №': a.inventory_number,
            'Наименование': a.name,
            'Подразделение': a.department or '',
            'Шифр': a.cipher or '',
            'Категория': a.category_rel.name if a.category_rel else '',
            'Стоимость': a.initial_cost,
            'Остаточная стоимость': a.salvage_value,
            'Срок пол.исп. (мес)': a.useful_life_months,
            'Дата ввода': a.start_date.isoformat() if a.start_date else '',
            'Место установки': a.location or '',
            'МОЛ': a.responsible_person or '',
            'Карточка': a.card or '',
            'Признак': a.indicator or '',
            'S/N': a.serial_number or '',
            'Тип': a.equipment_type or '',
            'Кол-во': a.quantity or 1,
            'Ед.изм.': a.unit or '',
            'Дата вывода (ЗИП)': a.inactive_date.isoformat() if a.inactive_date else '',
            'Дата работы (ЗИПа)': a.resume_date.isoformat() if a.resume_date else '',
            'Статус': a.status_display,
            'Дата статуса': a.status_date.isoformat() if a.status_date else '',
            'Примечания': a.notes or '',
        })
    sess.close()
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=card_name[:31])
        ws = writer.sheets[card_name[:31]]
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                v = str(cell.value) if cell.value else ''
                clen = sum(1.7 if ord(c) > 127 else 1 for c in v)
                max_len = max(max_len, clen)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 2, 42), 8)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
    buf.seek(0)
    safe = card_name.replace('/', '_').replace('\\', '_').replace(':', '_')
    return send_file(buf, download_name=f'{safe}.xlsx', as_attachment=True)


@app.route('/registry/<int:registry_id>/assets/<int:asset_id>/upload', methods=['POST'])
def asset_photo_upload(registry_id, asset_id):
    reg, sess = open_registry(registry_id)
    sess.get(Asset, asset_id) or abort(404)
    if 'photo' not in request.files:
        sess.close()
        return {'error': 'Нет файла'}, 400
    file = request.files['photo']
    if not file.filename:
        sess.close()
        return {'error': 'Файл не выбран'}, 400
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename)[1] or '.jpg'
    stored_name = f'{uuid.uuid4().hex}{ext}'
    file.save(os.path.join(UPLOAD_DIR, stored_name))
    photo = AssetPhoto(asset_id=asset_id, filename=stored_name, original_filename=file.filename)
    sess.add(photo)
    sess.commit()
    sess.close()
    return {'ok': True, 'filename': stored_name, 'original_name': file.filename}


@app.route('/registry/<int:registry_id>/delete', methods=['POST'])
def registry_delete(registry_id):
    reg = get_registry(registry_id)
    db_path = os.path.join(app_dir(), reg.db_file)
    db.session.delete(reg)
    db.session.commit()
    if os.path.exists(db_path):
        os.remove(db_path)
    flash(f'Реестр "{reg.name}" удалён', 'info')
    return redirect(url_for('index'))


@app.route('/registry/<int:registry_id>/assets/new', methods=['GET', 'POST'])
def asset_create(registry_id):
    reg, sess = open_registry(registry_id)
    categories = sess.query(Category).order_by(Category.name).all()
    if request.method == 'POST':
        asset = Asset(
            inventory_number=request.form['inventory_number'],
            name=request.form['name'],
            category_id=resolve_category(sess, request.form['category_name']),
            initial_cost=float(request.form['initial_cost']),
            salvage_value=float(request.form.get('salvage_value', 0)),
            useful_life_months=int(request.form['useful_life_months']),
            start_date=date.fromisoformat(request.form['start_date']),
            location=request.form.get('location', ''),
            responsible_person=request.form.get('responsible_person', ''),
            notes=request.form.get('notes', ''),
            department=request.form.get('department', ''),
            cipher=request.form.get('cipher', ''),
            serial_number=request.form.get('serial_number', ''),
            equipment_type=request.form.get('equipment_type', ''),
            quantity=int(request.form.get('quantity', 1)) or 1,
            unit=request.form.get('unit', ''),
            indicator=request.form.get('indicator', ''),
            card=request.form.get('card', ''),
        )
        d = request.form.get('inactive_date', '')
        asset.inactive_date = date.fromisoformat(d) if d else None
        d = request.form.get('resume_date', '')
        asset.resume_date = date.fromisoformat(d) if d else None
        d = request.form.get('status_date', '')
        asset.status_date = date.fromisoformat(d) if d else None
        indicator = request.form.get('indicator', '')
        key = indicator.strip().lower().replace('ё', 'е') if indicator else ''
        asset.status = AssetModel.INDICATOR_STATUS_MAP.get(key, 'undefined') if key else 'undefined'
        sess.add(asset)
        sess.commit()
        asset_id = asset.id
        sess.close()
        flash('Основное средство добавлено', 'success')
        return redirect(url_for('asset_detail', registry_id=registry_id, asset_id=asset_id))
    card_list = [r[0] for r in sess.query(Asset.card).filter(Asset.card.isnot(None), Asset.card != '').distinct().order_by(Asset.card).all()]
    html = render_template('asset_form.html', registry=reg, categories=categories,
                           asset=None, today=date.today().isoformat(), card_list=card_list)
    sess.close()
    return html


@app.route('/registry/<int:registry_id>/assets/<int:asset_id>')
def asset_detail(registry_id, asset_id):
    reg, sess = open_registry(registry_id)
    asset = sess.get(Asset, asset_id) or abort(404)
    schedule = []
    for month in range(1, asset.useful_life_months + 1):
        dep = asset.monthly_depreciation
        accumulated = round(month * dep, 2)
        residual = round(asset.initial_cost - accumulated, 2)
        schedule.append({'month': month, 'depreciation': dep,
                         'accumulated': accumulated, 'residual': residual})
    photos = sess.query(AssetPhoto).filter_by(asset_id=asset_id).order_by(AssetPhoto.uploaded_at.desc()).all()
    html = render_template('asset_detail.html', registry=reg, asset=asset,
                           depreciation_schedule=schedule, photos=photos,
                           status_logs=sess.query(StatusLog).filter_by(asset_id=asset_id).order_by(StatusLog.changed_at.desc()).all())
    sess.close()
    return html


@app.route('/registry/<int:registry_id>/assets/<int:asset_id>/edit', methods=['GET', 'POST'])
def asset_edit(registry_id, asset_id):
    reg, sess = open_registry(registry_id)
    asset = sess.get(Asset, asset_id) or abort(404)
    categories = sess.query(Category).order_by(Category.name).all()
    if request.method == 'POST':
        asset.inventory_number = request.form['inventory_number']
        asset.name = request.form['name']
        asset.category_id = resolve_category(sess, request.form['category_name'])
        asset.initial_cost = float(request.form['initial_cost'])
        asset.salvage_value = float(request.form.get('salvage_value', 0))
        asset.useful_life_months = int(request.form['useful_life_months'])
        asset.start_date = date.fromisoformat(request.form['start_date'])
        asset.location = request.form.get('location', '')
        asset.responsible_person = request.form.get('responsible_person', '')
        old_status = asset.status
        asset.notes = request.form.get('notes', '')
        asset.department = request.form.get('department', '')
        asset.cipher = request.form.get('cipher', '')
        asset.serial_number = request.form.get('serial_number', '')
        asset.equipment_type = request.form.get('equipment_type', '')
        asset.quantity = int(request.form.get('quantity', 1)) or 1
        asset.unit = request.form.get('unit', '')
        asset.indicator = request.form.get('indicator', '')
        asset.card = request.form.get('card', '')
        form_status = request.form.get('status', '')
        if form_status and form_status != old_status:
            asset.status = form_status
        elif asset.indicator:
            key = asset.indicator.strip().lower().replace('ё', 'е')
            asset.status = AssetModel.INDICATOR_STATUS_MAP.get(key, 'undefined')
        else:
            asset.status = 'undefined'
        d = request.form.get('inactive_date', '')
        asset.inactive_date = date.fromisoformat(d) if d else None
        d = request.form.get('resume_date', '')
        asset.resume_date = date.fromisoformat(d) if d else None
        d = request.form.get('status_date', '')
        asset.status_date = date.fromisoformat(d) if d else None
        if old_status != asset.status:
            sess.add(StatusLog(
                asset_id=asset.id, old_status=old_status,
                new_status=asset.status, changed_at=asset.status_date or date.today(),
                comment=request.form.get('status_comment', ''),
            ))
        sess.commit()
        asset_id_val = asset.id
        sess.close()
        flash('Основное средство обновлено', 'success')
        return redirect(url_for('asset_detail', registry_id=registry_id, asset_id=asset_id_val))
    card_list = [r[0] for r in sess.query(Asset.card).filter(Asset.card.isnot(None), Asset.card != '').distinct().order_by(Asset.card).all()]
    html = render_template('asset_form.html', registry=reg, categories=categories, asset=asset, today=date.today().isoformat(), card_list=card_list)
    sess.close()
    return html


@app.route('/registry/<int:registry_id>/assets/<int:asset_id>/delete', methods=['POST'])
def asset_delete(registry_id, asset_id):
    reason = request.form.get('reason', '').strip()
    reg, sess = open_registry(registry_id)
    asset = sess.get(Asset, asset_id) or abort(404)
    asset.is_deleted = True
    asset.deleted_reason = reason or None
    asset.deleted_at = date.today()
    sess.commit()
    sess.close()
    flash('Основное средство перемещено в корзину', 'info')
    return redirect(url_for('registry_view', registry_id=registry_id))


@app.route('/registry/<int:registry_id>/trash')
def trash_view(registry_id):
    reg, sess = open_registry(registry_id)
    assets = sess.query(Asset).filter(Asset.is_deleted == True).order_by(Asset.deleted_at.desc()).all()
    sess.close()
    return render_template('trash.html', registry=reg, assets=assets)


@app.route('/registry/<int:registry_id>/trash/<int:asset_id>/restore', methods=['POST'])
def trash_restore(registry_id, asset_id):
    reg, sess = open_registry(registry_id)
    asset = sess.get(Asset, asset_id) or abort(404)
    asset.is_deleted = False
    asset.deleted_reason = None
    asset.deleted_at = None
    sess.commit()
    sess.close()
    flash('Основное средство восстановлено', 'success')
    return redirect(url_for('trash_view', registry_id=registry_id))


@app.route('/registry/<int:registry_id>/trash/<int:asset_id>/destroy', methods=['POST'])
def trash_destroy(registry_id, asset_id):
    reg, sess = open_registry(registry_id)
    asset = sess.get(Asset, asset_id) or abort(404)
    for photo in sess.query(AssetPhoto).filter_by(asset_id=asset_id).all():
        p = os.path.join(app_dir(), 'uploads', 'photos', photo.filename)
        if os.path.exists(p):
            os.remove(p)
        sess.delete(photo)
    sess.delete(asset)
    sess.commit()
    sess.close()
    flash('Основное средство удалено навсегда', 'info')
    return redirect(url_for('trash_view', registry_id=registry_id))


@app.route('/registry/<int:registry_id>/assets/<int:asset_id>/movement', methods=['GET', 'POST'])
def movement_create(registry_id, asset_id):
    reg, sess = open_registry(registry_id)
    asset = sess.get(Asset, asset_id) or abort(404)
    if request.method == 'POST':
        movement = Movement(
            asset_id=asset.id, date=date.fromisoformat(request.form['date']),
            movement_type=request.form['movement_type'],
            from_location=request.form.get('from_location', ''),
            to_location=request.form.get('to_location', ''),
            reason=request.form.get('reason', ''),
            document_number=request.form.get('document_number', ''),
        )
        if movement.movement_type == 'disposal':
            asset.status = 'disposed'
        elif movement.movement_type == 'transfer':
            asset.status = 'transferred'
            if movement.to_location:
                asset.location = movement.to_location
        sess.add(movement)
        sess.commit()
        asset_id_val = asset.id
        sess.close()
        flash('Перемещение зарегистрировано', 'success')
        return redirect(url_for('asset_detail', registry_id=registry_id, asset_id=asset_id_val))
    html = render_template('movement_form.html', registry=reg, asset=asset,
                           today=date.today().isoformat())
    sess.close()
    return html


# ─── Import from Excel ─────────────────────────────────────────────────

@app.route('/registry/<int:registry_id>/import', methods=['GET', 'POST'])
def asset_import(registry_id):
    reg, sess = open_registry(registry_id)
    categories = {c.name.lower(): c for c in sess.query(Category).all()}

    if request.method == 'POST':
        if 'file' not in request.files or not request.files['file'].filename:
            flash('Файл не выбран', 'danger')
            sess.close()
            return render_template('import_form.html', registry=reg)

        file = request.files['file']
        try:
            file_bytes = io.BytesIO(file.read())
            raw = pd.read_excel(file_bytes, header=None, dtype=str)
        except Exception as e:
            flash(f'Ошибка чтения Excel: {e}', 'danger')
            sess.close()
            return render_template('import_form.html', registry=reg)

        # Detect header row
        header_row = 0
        if len(raw) > 1:
            row0 = [normalize(str(v)) for v in raw.iloc[0] if pd.notna(v)]
            row1 = [normalize(str(v)) for v in raw.iloc[1] if pd.notna(v)]
            def score(vals):
                return sum(1 for v in vals if any(nk in v or v in nk for nk in FIELD_MAP))
            if score(row1) > score(row0):
                header_row = 1

        if len(raw) <= header_row:
            flash('Файл пуст', 'danger')
            sess.close()
            return render_template('import_form.html', registry=reg)

        file_bytes.seek(0)
        df = pd.read_excel(file_bytes, header=header_row, dtype=str)

        col_map = guess_columns(df.columns)
        if 'inventory_number' not in col_map or 'name' not in col_map:
            flash('Не найдены обязательные колонки: Инв. №, Наименование', 'warning')
            sess.close()
            return render_template('import_form.html', registry=reg)

        imported, errors = 0, 0
        for idx, row in df.iterrows():
            try:
                inv = str(row.iloc[col_map['inventory_number']]).strip() if pd.notna(row.iloc[col_map['inventory_number']]) else ''
                nm  = str(row.iloc[col_map['name']]).strip() if pd.notna(row.iloc[col_map['name']]) else ''
                if not inv or not nm:
                    errors += 1
                    continue

                cat_name = ''
                if 'category' in col_map:
                    cat_name = str(row.iloc[col_map['category']]) if pd.notna(row.iloc[col_map['category']]) else ''
                cat = categories.get(cat_name.strip().lower()) if cat_name else None
                if not cat:
                    cat = categories.get('прочие основные средства') or list(categories.values())[0]

                asset = Asset(
                    inventory_number=inv, name=nm, category_id=cat.id,
                    initial_cost=0.0, salvage_value=0.0, useful_life_months=60,
                    start_date=date.today(),
                )
                for field_name, col_idx in col_map.items():
                    val = row.iloc[col_idx]
                    if field_name == 'initial_cost':
                        asset.initial_cost = parse_float(val)
                    elif field_name == 'salvage_value':
                        asset.salvage_value = parse_float(val)
                    elif field_name == 'useful_life_months':
                        asset.useful_life_months = int(parse_float(val)) if parse_float(val) > 0 else 60
                    elif field_name == 'start_date':
                        asset.start_date = parse_date_val(val)
                    elif field_name in ('location', 'responsible_person', 'notes', 'department',
                                   'cipher', 'serial_number', 'equipment_type', 'unit', 'indicator', 'info', 'card'):
                        setattr(asset, field_name, str(val) if pd.notna(val) else '')
                    elif field_name == 'quantity':
                        asset.quantity = int(parse_float(val)) if parse_float(val) > 0 else 1
                    elif field_name == 'inactive_date':
                        asset.inactive_date = parse_date_val(val) if pd.notna(val) else None
                    elif field_name == 'resume_date':
                        asset.resume_date = parse_date_val(val) if pd.notna(val) else None

                # Auto-determine status from indicator
                key = asset.indicator.strip().lower().replace('ё', 'е') if asset.indicator else ''
                if key:
                    asset.status = AssetModel.INDICATOR_STATUS_MAP.get(key, 'undefined')
                else:
                    asset.status = 'undefined'
                asset.status_date = None

                existing = sess.query(Asset).filter_by(inventory_number=inv).first()
                if existing:
                    for col in Asset.__table__.columns:
                        if col.name != 'id':
                            setattr(existing, col.name, getattr(asset, col.name))
                    sess.flush()
                else:
                    with sess.begin_nested():
                        sess.add(asset)
                imported += 1
            except Exception:
                errors += 1

        sess.commit()
        sess.close()
        msg = f'Импортировано: {imported}'
        if errors:
            msg += f', пропущено: {errors}'
        flash(msg, 'success')
        return redirect(url_for('registry_view', registry_id=registry_id))

    sess.close()
    return render_template('import_form.html', registry=reg)


# ─── Init & Run ────────────────────────────────────────────────────────

if __name__ == '__main__':
    os.makedirs(os.path.join(app_dir(), 'instance'), exist_ok=True)
    os.makedirs(os.path.join(app_dir(), 'databases'), exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    with app.app_context():
        db.create_all()
    app.run(debug=False, host='0.0.0.0', port=5000)
